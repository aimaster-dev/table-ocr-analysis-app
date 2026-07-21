"""Smoke tests for pure helpers (no OCR / Qt required)."""

from __future__ import annotations

from table_scan.core.excel_writer import ExcelExporter
from table_scan.core.html_writer import HtmlExporter
from table_scan.models.table_result import ExtractedTable
from html.parser import HTMLParser
from openpyxl import load_workbook


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        del attrs
        if tag == "tr":
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def test_extracted_table_pads_ragged_rows() -> None:
    table = ExtractedTable(rows=[["a", "b"], ["c"], ["d", "e", "f"]])
    rect = table.to_rectangular()
    assert rect == [["a", "b", ""], ["c", "", ""], ["d", "e", "f"]]


def test_excel_exporter_writes_file(tmp_path) -> None:
    table = ExtractedTable(rows=[["Name", "Qty"], ["Apple", "3"], ["Pear", "1"]])
    out = tmp_path / "sample.xlsx"
    ExcelExporter().export([table], out)
    assert out.is_file()
    assert out.stat().st_size > 0


def test_excel_exporter_preserves_detected_merged_header(tmp_path) -> None:
    table = ExtractedTable(
        rows=[["販売報告", "", ""], ["商品", "数量", "Price"]],
        merged_ranges=[(0, 0, 0, 2)],
    )
    out = tmp_path / "merged.xlsx"
    ExcelExporter().export([table], out)

    sheet = load_workbook(out).active
    assert str(next(iter(sheet.merged_cells.ranges))) == "A1:C1"
    assert sheet["A1"].value == "販売報告"


def test_excel_exporter_flags_low_confidence_handwriting(tmp_path) -> None:
    table = ExtractedTable(
        rows=[["항목", "Value"], ["손글씨", "12"]],
        confidences=[[0.98, 0.97], [0.51, 0.94]],
    )
    out = tmp_path / "review.xlsx"
    ExcelExporter().export([table], out)

    sheet = load_workbook(out).active
    assert sheet["A2"].fill.fill_type == "solid"
    assert sheet["A2"].fill.fgColor.rgb == "00FFF2CC"
    assert sheet["B2"].fill.fill_type is None


def test_excel_exporter_expands_multiline_rows_and_hides_default_grid(tmp_path) -> None:
    table = ExtractedTable(
        rows=[
            ["Topic", "Key Points"],
            ["Time", "Prioritize tasks\nUse calendar\nAvoid multitasking"],
        ]
    )
    out = tmp_path / "multiline.xlsx"
    ExcelExporter().export([table], out)

    sheet = load_workbook(out).active
    assert sheet.sheet_view.showGridLines is False
    assert sheet.row_dimensions[2].height >= 48


def test_html_exporter_escapes_ocr_text_and_preserves_merges(tmp_path) -> None:
    table = ExtractedTable(
        rows=[
            ["<script>alert('x')</script>", "", "Qty"],
            ["Tea & coffee", "", "2"],
        ],
        confidences=[[0.52, 0.0, 0.99], [0.98, 0.0, 0.97]],
        merged_ranges=[(0, 0, 0, 1), (1, 0, 1, 1)],
    )
    out = tmp_path / "safe.html"

    HtmlExporter().export([table], out, title="A < B")
    markup = out.read_text(encoding="utf-8")

    assert "<script>alert" not in markup
    assert "&lt;script&gt;alert('x')&lt;/script&gt;" in markup
    assert "A &lt; B" in markup
    assert 'colspan="2"' in markup
    assert 'class="review"' in markup
    assert "Content-Security-Policy" in markup


def test_html_and_excel_exporters_preserve_the_same_cell_matrix(tmp_path) -> None:
    table = ExtractedTable(
        rows=[
            ["Topic", "Key Points", "Notes"],
            ["1. Time\nManagement", "• First\n• Second", "한국어 & 中文"],
        ],
        confidences=[[0.99, 0.99, 0.99], [0.95, 0.94, 0.93]],
    )
    excel_path = tmp_path / "parity.xlsx"
    html_path = tmp_path / "parity.html"
    ExcelExporter().export([table], excel_path)
    HtmlExporter().export([table], html_path, title="Parity")

    sheet = load_workbook(excel_path).active
    excel_values = [
        [sheet.cell(row=row, column=column).value or "" for column in range(1, 4)]
        for row in range(1, 3)
    ]
    parser = _HtmlTableParser()
    parser.feed(html_path.read_text(encoding="utf-8"))

    assert excel_values == table.to_rectangular()
    assert parser.rows == table.to_rectangular()
