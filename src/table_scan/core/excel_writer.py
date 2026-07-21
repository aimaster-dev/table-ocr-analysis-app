"""Write extracted tables to Excel workbooks."""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from table_scan.models.table_result import ExtractedTable


class ExcelExporter:
    """Create formatted .xlsx files from extracted table matrices."""

    def __init__(
        self,
        *,
        bold_header: bool = True,
        review_confidence: float = 0.65,
    ) -> None:
        self.bold_header = bold_header
        self.review_confidence = review_confidence

    def export(
        self,
        tables: list[ExtractedTable],
        output_path: Path,
        *,
        sheet_prefix: str = "Table",
    ) -> Path:
        if not tables:
            raise ValueError("No tables to export")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        # Remove the default sheet; we recreate per table.
        default = workbook.active
        workbook.remove(default)

        thin = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        header_font = Font(bold=True)
        align = Alignment(vertical="center", wrap_text=True)
        review_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

        for index, table in enumerate(tables, start=1):
            title = f"{sheet_prefix}{index}" if len(tables) > 1 else sheet_prefix
            sheet = workbook.create_sheet(title=self._safe_sheet_name(title))
            sheet.sheet_view.showGridLines = False
            matrix = table.to_rectangular()

            for r_idx, row in enumerate(matrix, start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin
                    cell.alignment = align
                    if self.bold_header and r_idx == 1:
                        cell.font = header_font
                    if (
                        value
                        and r_idx - 1 < len(table.confidences)
                        and c_idx - 1 < len(table.confidences[r_idx - 1])
                    ):
                        confidence = table.confidences[r_idx - 1][c_idx - 1]
                        if 0.0 < confidence < self.review_confidence:
                            cell.fill = review_fill

            for r1, c1, r2, c2 in table.merged_ranges:
                if r2 >= r1 and c2 > c1:
                    sheet.merge_cells(
                        start_row=r1 + 1,
                        start_column=c1 + 1,
                        end_row=r2 + 1,
                        end_column=c2 + 1,
                    )

            self._autosize(sheet, matrix)

        workbook.save(output_path)
        return output_path

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        invalid = set(r"[]:*?/\\")
        cleaned = "".join("_" if ch in invalid else ch for ch in name).strip() or "Sheet"
        return cleaned[:31]

    @staticmethod
    def _autosize(sheet, matrix: list[list[str]]) -> None:
        if not matrix:
            return
        col_count = len(matrix[0])
        widths: list[float] = []
        for col in range(1, col_count + 1):
            max_len = 0
            for row in matrix:
                text = row[col - 1] if col - 1 < len(row) else ""
                lines = str(text).splitlines() or [""]
                max_len = max(max_len, *(len(line) for line in lines))
            width = min(max(max_len + 2, 10), 48)
            widths.append(width)
            sheet.column_dimensions[get_column_letter(col)].width = width

        # Excel does not automatically expand wrapped rows in every viewer.
        # Estimate the necessary height from explicit newlines and wrapping so
        # multi-line OCR results are visible when the workbook first opens.
        for row_index, row in enumerate(matrix, start=1):
            line_count = 1
            for col_index, value in enumerate(row):
                text_lines = str(value or "").splitlines() or [""]
                usable = max(1, int(widths[col_index] - 1))
                wrapped = sum(max(1, math.ceil(len(line) / usable)) for line in text_lines)
                line_count = max(line_count, wrapped)
            sheet.row_dimensions[row_index].height = min(120, max(18, 15 * line_count + 3))
